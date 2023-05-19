import openpyxl as xl
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox, ttk


def submit():
    file = xl.load_workbook('Data.xlsx')
    sheet = file.active
    product_value = product_entry.get()
    category_value = category_entry.get()
    color_value = color_entry.get()

    sheet.cell(column=1, row=sheet.max_row + 1, value=product_value)
    sheet.cell(column=2, row=sheet.max_row, value=category_value)
    sheet.cell(column=3, row=sheet.max_row, value=color_value)

    file.save('Data.xlsx')
    messagebox.showinfo(title="Success", message="Data has been saved")


def clear():
    product_entry.delete(0, END)
    color_entry.delete(0, END)


window = Tk()
window.title("Data Entry Form")
window.geometry("500x300")
window.config(bg="#290514")

font1 = ("Arial", 20, 'bold')
font2 = ("Arial", 15, 'bold')

product_label = Label(window, text="Product:", font=font1, width=10)
product_label.place(x=20, y=30)

category_label = Label(window, text="Category:", font=font1, width=10)
category_label.place(x=20, y=90)

color_label = Label(window, text="Color:", font=font1, width=10)
color_label.place(x=20, y=150)

product = StringVar()
color = StringVar()
category = StringVar()

product_entry = Entry(window, textvariable=product, font=font2, width=25)
product_entry.place(x=190, y=33)

color_entry = Entry(window, textvariable=color, font=font2, width=25)
color_entry.place(x=190, y=155)

category_entry = ttk.Combobox(window, font=font2, values=['PC', 'Smartphone'], width=24)
category_entry.place(x=190, y=95)

submit_button = Button(window, text="Submit", font=font1, fg="#0a8018", command=submit)
submit_button.place(x=190, y=210)

clear_button = Button(window, text="Clear", font=font1, command=clear)
clear_button.place(x=330, y=210)

window.mainloop()
